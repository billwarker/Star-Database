import openpyxl
import os
from bs4 import BeautifulSoup
import datetime
from Files.file_io import LeanDownloader, _convert_xml_to_xlsx
from settings import DOWNLOAD_DIR, STAR_PAYLOAD, SBW_PAYLOAD, CHROMEDRIVER, INVENTORY_DIR
from settings import psql

from sqlalchemy import create_engine, exists
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from database import StarInventory, SBWInventory

class InventoryUpdater:
	def __init__(self, psql):
		self.engine = create_engine('postgresql://{}:{}@{}:{}/{}'.format(
											psql['user'], psql['password'],
											psql['host'], psql['port'],
											psql['database']))
		self.Base = declarative_base()
		Session = sessionmaker(bind=self.engine)
		self.session = Session()

	def update_star(self, inventory_file):

		print('Updating inventory for Star Interactive...')

		wb = openpyxl.load_workbook(inventory_file)
		sheet = wb.active

		added_skus = set()
		added_count = 0
		updated_skus = set()

		for row in range(2, sheet.max_row + 1):

			item = StarInventory(item_num = sheet['A' + str(row)].value,
								item_sku = sheet['B' + str(row)].value,
								item_desc = sheet['C' + str(row)].value,
								item_inv = int(sheet['E' + str(row)].value),
								item_upc = sheet['I' + str(row)].value)

			in_database = self.session.query(exists().\
							where(StarInventory.item_num == item.item_num)).\
							first()[0]
			if in_database:
				prev = self.session.query(StarInventory.item_inv).\
							filter(StarInventory.item_num == item.item_num).first()[0]
				if prev != item.item_inv:
					self.session.query(StarInventory).\
							filter(StarInventory.item_num == item.item_num).\
							update({StarInventory.item_inv: item.item_inv})

					updated_skus.add("{} {} ---> {}".format(item.item_sku, prev, item.item_inv))
					self.session.commit()
			else:
				self.session.add(item)
				self.session.commit()
				added_skus.add("{} ({})".format(item.item_sku, item.item_inv))

		print('SKUs added:', len(added_skus))
		for sku in added_skus:
			print(sku)
		print("---")
		print('SKUs updated:', len(updated_skus))
		for sku in updated_skus:
			print(sku)
	
	def update_sbw(self, inventory_file):
		
		print('Updating inventory for SBW...')

		wb = openpyxl.load_workbook(inventory_file)
		sheet = wb.active

		added_skus = set()
		added_count = 0
		updated_skus = set()

		for row in range(2, sheet.max_row + 1):

			item = SBWInventory(item_num = sheet['A' + str(row)].value,
								item_sku = sheet['B' + str(row)].value,
								item_desc = sheet['C' + str(row)].value,
								item_inv = int(sheet['E' + str(row)].value),
								item_upc = sheet['I' + str(row)].value)

			in_database = self.session.query(exists().\
							where(SBWInventory.item_num == item.item_num)).\
							first()[0]
			if in_database:
				prev = self.session.query(SBWInventory.item_inv).\
							filter(SBWInventory.item_num == item.item_num).first()[0]
				if prev != item.item_inv:
					self.session.query(SBWInventory).\
							filter(SBWInventory.item_num == item.item_num).\
							update({SBWInventory.item_inv: item.item_inv})

					updated_skus.add("{} {} ---> {}".format(item.item_sku, prev, item.item_inv))
					self.session.commit()
			else:
				self.session.add(item)
				self.session.commit()
				added_skus.add("{} ({})".format(item.item_sku, item.item_inv))

		print('SKUs added:', len(added_skus))
		for sku in added_skus:
			print(sku)
		print("---")
		print('SKUs updated:', len(updated_skus))
		for sku in updated_skus:
			print(sku)

if __name__ == '__main__':
	os.makedirs(INVENTORY_DIR, exist_ok=True)
	
	name1 = "STAR Inventory {}".format(datetime.date.today().strftime("%m-%d-%Y"))
	name2 = "SBW Inventory {}".format(datetime.date.today().strftime("%m-%d-%Y"))
	
	downloader = LeanDownloader(CHROMEDRIVER)
	downloader.login(STAR_PAYLOAD)
	star_inventory = _convert_xml_to_xlsx(downloader.download_inventory(DOWNLOAD_DIR), name1, INVENTORY_DIR)

	downloader.logout()
	downloader.login(SBW_PAYLOAD)
	sbw_inventory = _convert_xml_to_xlsx(downloader.download_inventory(DOWNLOAD_DIR), name2, INVENTORY_DIR)
	downloader.close()

	updater = InventoryUpdater(psql)
	updater.update_star(star_inventory)
	updater.update_sbw(sbw_inventory)