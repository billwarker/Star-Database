import openpyxl
import sqlalchemy
from sqlalchemy import create_engine, exists
from sqlalchemy.orm import sessionmaker
from database import AmazonOrder
from sqlalchemy.ext.declarative import declarative_base
from models import LeanOrder
from datetime import timedelta, date
from settings import psql
import datetime

def find_retailer(client_order, end_client_order_no):
	try:
		end_client_order_no = str(end_client_order_no)
		client_order = str(client_order)

		if end_client_order_no.startswith("Y") or \
			client_order.startswith("Y"):
			return "Walmart Canada"
		elif end_client_order_no.startswith("4") or end_client_order_no.startswith("6") or \
			client_order.startswith("4") or client_order.startswith("4"):
			return "Best Buy Canada"
		elif end_client_order_no.startswith("H") or client_order.startswith("H"):
			return "The Source"
		elif end_client_order_no.startswith("5") or client_order.startswith("5"):
			return "Staples"
		elif end_client_order_no.startswith("2") or client_order.startswith("2"):
			return "Groupon"
		if end_client_order_no.startswith("3") or client_order.startswith("3"):
			return "Shop.ca"
		else:
			return None

	except AttributeError:
		return None

def location(destination, province=False, city=False):
	try:
		location = destination.split('@')[1]
		if province:
			return location.split(' ')[-1]
		elif city:
			return location.replace(location.split(' ')[-1], '')
		else:
			return None
	except AttributeError:
		return None
	except IndexError:
		return destination

def find_customer_name(destination):
	try:
		return destination.split('@')[0]
	except AttributeError:
		return None
	except IndexError:
		return destination

class OrderParser:
	def __init__(self, psql):
		self.engine = create_engine('postgresql://{}:{}@{}:{}/{}'.format(
											psql['user'], psql['password'],
											psql['host'], psql['port'],
											psql['database']))
		self.Base = declarative_base()
		Session = sessionmaker(bind=self.engine)
		self.session = Session()

	def add_everything(self, order_data):
		print("Adding all Lean orders to database...")
		self.order_data = order_data
		self.input_wb = openpyxl.load_workbook(self.order_data)
		self.input_sheet = self.input_wb.active

		for row in range(2, self.input_sheet.max_row + 1):
			# This info is used to find retailer, prov, city, customer name
			client_order = self.input_sheet["B" + str(row)].value
			destination = self.input_sheet["D" + str(row)].value
			end_client_order_no = self.input_sheet["C" + str(row)].value

			order = LeanOrder(order_id = self.input_sheet["A" + str(row)].value,
								end_client_order_no = end_client_order_no,
								destination = destination,
								created_on = self.input_sheet["E" + str(row)].value,
								order_status = self.input_sheet["F" + str(row)].value,
								ship_date =self. input_sheet["I" + str(row)].value,
								item = self.input_sheet["J" + str(row)].value,
								upc = self.input_sheet["K" + str(row)].value,
								qty = self.input_sheet["M" + str(row)].value,
								ship_qty = self.input_sheet["M" + str(row)].value,
								carrier = self.input_sheet["O" + str(row)].value,
								tracking_no = self.input_sheet["P" + str(row)].value,
								#shipment_no = self.input_sheet["Q" + str(row)].value,
								retailer = find_retailer(client_order, end_client_order_no),
								province = location(destination, province=True),
								city = location(destination, city=True),
								customer_name = find_customer_name(destination))
			
			self.session.add(order)
			print("Added order {}".format(end_client_order_no))
		self.session.commit()
		print("Orders committed.")

	def update(self, order_data):
		print("Updating Lean Orders...")
		self.order_data = order_data
		self.input_wb = openpyxl.load_workbook(self.order_data)
		self.input_sheet = self.input_wb.active

		for row in range(self.input_sheet.max_row - 250, self.input_sheet.max_row + 1):
			# This info is used to find retailer, prov, city, customer name
			client_order = self.input_sheet["B" + str(row)].value
			destination = self.input_sheet["D" + str(row)].value
			end_client_order_no = self.input_sheet["C" + str(row)].value
			
			order = LeanOrder(order_id = self.input_sheet["A" + str(row)].value,
								end_client_order_no = end_client_order_no,
								destination = destination,
								created_on = self.input_sheet["E" + str(row)].value,
								order_status = self.input_sheet["F" + str(row)].value,
								ship_date =self. input_sheet["I" + str(row)].value,
								item = self.input_sheet["J" + str(row)].value,
								upc = self.input_sheet["K" + str(row)].value,
								qty = self.input_sheet["M" + str(row)].value,
								ship_qty = self.input_sheet["M" + str(row)].value,
								carrier = self.input_sheet["O" + str(row)].value,
								tracking_no = self.input_sheet["P" + str(row)].value,
								#shipment_no = self.input_sheet["Q" + str(row)].value,
								retailer = find_retailer(client_order, end_client_order_no),
								province = location(destination, province=True),
								city = location(destination, city=True),
								customer_name = find_customer_name(destination))

			in_database = self.session.query(exists().\
							where(LeanOrder.order_id == order.order_id)).\
							first()[0]
			if not in_database:
				self.session.add(order)
				print("Added order {}".format(order.end_client_order_no))

		self.session.commit()
		print("Orders committed.")

if __name__ == '__main__':
	parser = OrderParser(psql)
	parser.add_everything("OrderDetailData (27).xlsx")
	parser.update("OrderDetailData (34).xlsx")